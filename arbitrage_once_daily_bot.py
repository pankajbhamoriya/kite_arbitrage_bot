import time
from datetime import datetime, timedelta
from kiteconnect import KiteConnect, KiteTicker
from openpyxl import Workbook, load_workbook

API_KEY = "your_kite_api_key"
ACCESS_TOKEN = "your_access_token"
NIFTY_FUT_TOKEN = 123456  # Replace with actual token
BANKNIFTY_FUT_TOKEN = 654321  # Replace with actual token
EXCEL_FILE = "arbitrage_orders.xlsx"

kite = KiteConnect(api_key=API_KEY)
kite.set_access_token(ACCESS_TOKEN)

def init_excel():
    try:
        wb = load_workbook(EXCEL_FILE)
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.append(["Timestamp", "Action", "Instrument", "Price", "P&L"])
        wb.save(EXCEL_FILE)

def log_to_excel(action, instrument, price, pnl=""):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), action, instrument, price, pnl])
    wb.save(EXCEL_FILE)

class DailyArbitrageBot:
    def __init__(self):
        self.start_prices = {}
        self.last_prices = {}
        self.order_placed_today = False
        self.order_time = None
        self.positions = {}
        self.ticker = KiteTicker(API_KEY, ACCESS_TOKEN)
        self.ticker.on_ticks = self.on_ticks
        self.ticker.on_connect = self.on_connect
        self.ticker.on_close = self.on_close

    def start(self):
        print("Starting KiteTicker...")
        self.ticker.connect(threaded=True, disable_ssl_verification=True)
        while True:
            now = datetime.now()
            if self.order_placed_today and self.order_time:
                # Check every hour after order placed
                if (now - self.order_time).seconds >= 3600:
                    self.check_pnl()
                    self.order_time = now  # Reset timer for next hourly check
            if now.hour == 9 and now.minute == 15 and not self.order_placed_today:
                # Reset daily order flag at market open (customize as per market timings)
                self.order_placed_today = False
                self.positions = {}
                self.start_prices = {}
                print("Resetting for new day...")
            time.sleep(30)

    def on_connect(self, ws, response):
        ws.subscribe([NIFTY_FUT_TOKEN, BANKNIFTY_FUT_TOKEN])
        ws.set_mode(ws.MODE_FULL, [NIFTY_FUT_TOKEN, BANKNIFTY_FUT_TOKEN])

    def on_ticks(self, ws, ticks):
        for tick in ticks:
            token = tick['instrument_token']
            ltp = tick['last_price']
            if token not in self.start_prices:
                self.start_prices[token] = ltp
            self.last_prices[token] = ltp

        if len(self.last_prices) == 2 and not self.order_placed_today:
            self.check_arbitrage()

    def on_close(self, ws, code, reason):
        print("WebSocket closed:", reason)

    def check_arbitrage(self):
        nifty_open = self.start_prices[NIFTY_FUT_TOKEN]
        banknifty_open = self.start_prices[BANKNIFTY_FUT_TOKEN]
        nifty_ltp = self.last_prices[NIFTY_FUT_TOKEN]
        banknifty_ltp = self.last_prices[BANKNIFTY_FUT_TOKEN]

        pct_nifty = ((nifty_ltp - nifty_open) / nifty_open) * 100
        pct_banknifty = ((banknifty_ltp - banknifty_open) / banknifty_open) * 100
        diff = abs(pct_nifty - pct_banknifty)

        if diff > 0.5:
            # Place arbitrage trade only once per day
            if pct_nifty > pct_banknifty:
                self.place_order("SELL", "NIFTY", nifty_ltp)
                self.place_order("BUY", "BANKNIFTY", banknifty_ltp)
                self.positions = {"NIFTY": nifty_ltp, "BANKNIFTY": banknifty_ltp}
            else:
                self.place_order("SELL", "BANKNIFTY", banknifty_ltp)
                self.place_order("BUY", "NIFTY", nifty_ltp)
                self.positions = {"NIFTY": nifty_ltp, "BANKNIFTY": banknifty_ltp}
            self.order_placed_today = True
            self.order_time = datetime.now()

    def place_order(self, action, instrument, price):
        # Add your kite.place_order() logic here
        print(f"{action} {instrument} at {price}")
        log_to_excel(action, instrument, price)

    def check_pnl(self):
        print("Checking P&L...")
        # Fetch current prices
        ltp_nifty = self.last_prices.get(NIFTY_FUT_TOKEN)
        ltp_banknifty = self.last_prices.get(BANKNIFTY_FUT_TOKEN)
        pnl = 0.0

        if self.positions:
            # Calculate P&L for both legs
            for inst, entry_price in self.positions.items():
                if inst == "NIFTY":
                    curr_price = ltp_nifty
                else:
                    curr_price = ltp_banknifty
                if curr_price:
                    if entry_price < curr_price:
                        pnl_leg = curr_price - entry_price
                        action = "BUY"
                    else:
                        pnl_leg = entry_price - curr_price
                        action = "SELL"
                    log_to_excel("PnL_Check", inst, curr_price, pnl_leg)
                    pnl += pnl_leg

            print(f"Total P&L: {pnl}")

if __name__ == "__main__":
    init_excel()
    bot = DailyArbitrageBot()
    bot.start()